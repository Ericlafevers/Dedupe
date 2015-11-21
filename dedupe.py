import openpyxl as pxl
from glob import glob

def find_files(files):
    all_files = []
    for fname in glob('*.xlsx'):
        all_files.append(fname)
        return all_files

def select_master(all_files, choice):
    print ('Please select your master file: ')
    for f in all_files:
        print (all_files.index(f), ' - ', f)
        
    file_choice = raw_input('Choose your file with the corresponding number: ')
    if file_choice != int(): #type(file_choice) != int()
        print ('That is not a valid selection, try again')
    if file_choice == int():
        master_file = file_choice
        all_files = all_files.remove(master_file)
        return master_file, all_files
        
def read_files(files, all_files, master_file):
    master_read = 0
    rest_read = 1
    
    master_data = []
    rest_data = []
    
    while master_read == 0:
        masterWB = pxl.load_workbook(master_file, data_only=True)
        masterSheet = masterWB.active
        for r in masterSheet.iter_rows():
            master_data.append(r)
            #for c in r:
        master_read = 1
        rest_read = 0
        
    while rest_read == 0:
        for f in all_files:
            restWB = pxl.load_workbook(f, data_only=True)
            restSheet = restWB.active
            for r in restSheet.iter_rows():
                rest_data.append(r)
            rest_read = 1

def compare_data(master_data, rest_data):
    dupe_phone = []
    dupe_email []
    
    for i in rest_data:
        for v in i:
            if v in master_data:
                if '@' in v:
                    dupe_email.append(i)
                elif type(v) == int(): # if int() in v:
                    dupe_phone.append(i)
                else:
                    continue

def get_missing(master_data, dupe_phone, dupe_email):
    missing_data = []
    
    for i in master_data:
        for v in i:
            if v in dupe_phone or dupe_email:
                missing_data.appened(i)
                
def write_files(savenames, missing_savename, dupe_phone_savename, dupe_email_savename):
    for f in savenames:
        newWB = pxl.Workbook()
        sheet = newWB.active
        
        
        if f == missing_savename:
            for i in missing_data:
                rowcount = missing_data.index(i) + 1
                #colcount = missing_data.index(i[v]) + 1
                for v in i:
                    colcount = i.index(v) + 1
                    sheet.cell(row = rowcount, column = colcount).value = v
        if f == dupe_phone_savename:
            for i in dupe_phone:
                rowcount = dupe_phone.index(i) + 1
                for v in i:
                    colcount = i.index(v) + 1
                    sheet.cell(row = rowcount, column = colcount).value = v
        if f == dupe_email_savename:
            for i in dupe_email:
                rowcount = dupe_email.index(i) + 1
                for v in i:
                    colcount = i.index(v) + 1
                    sheet.cell(row = rowcount, column = colcount).value = v
        
        newWB.save(f)
    
if __name__ == '__main__':
    
    missing_savename = 'Missing_data.xlsx'
    dupe_phone_savename = 'Duplicate_phone.xlsx'
    dupe_email_savename = 'Duplicate_email.xlsx'
    savenames = [missing_savename, dupe_phone_savename, dupe_email_savename]
    
    # TODO figure out how to call all functions flawlessly - area of practice for me
    
    
    input('\n\nPress ENTER to quite.')
    
    
    
    
    